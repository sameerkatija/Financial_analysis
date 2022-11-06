import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import islice
from GrowthRatios import *
from StabilityRatios import * 
# Path to the file
file_path = 'FA.xlsx'

# import excel file
dataFile = load_workbook(file_path)

# Object of the sheet 
dataSheet = dataFile.active

dataFrame = dataSheet.values

# print(dataFrame)
cols = next(dataFrame)[1:]
dataFrame = list(dataFrame)
idx = [r[0] for r in dataFrame]
dataFrame = (islice(r, 1, None) for r in dataFrame)
df = pd.DataFrame(dataFrame, index=idx, columns=cols)

new_df = pd.concat({
    "Sales Growth": SalesGrowth(df),
    "EPS": EPSGrowth(df),
    "Net Profit Growth": NetProfitGrowth(df),
    "Operating Profit Growth": OperatingProfitGrowth(df),
    "OPM": OPM(df),
    "NPM": NPM(df),
    "Tax Ratio": TaxRatio(df),
    "Interest Coverage":InterestCoverage(df),
    "Total Debt": TotalDebt(df),
    "Debt To Equity": DebtToEquity(df),
    "Current Ratio":CurrentRatio(df),
    "Cash Flow From Operating Activities": CFO(df),
    "Net Change in Cash": NetCash(df),
    "cCFO vs cPAT": CCFOVSPAT(df),
    "Net Asset Turnover":NetAssetTurnover(df), 
    "Return on Equity": ReturnOnEquity(df)
}, axis=1).reset_index()

# # ratiosDF
Ratio = pd.DataFrame(new_df).set_index('index')
Ratio.index.name = 'Ratios'
Ratio.T.to_excel('check.xlsx')
# print(Ratio.T)


# ratiosDf = pd.DataFrame(ratiosDf)
# ratiosDF.to_excel('check.xlsx')

# wb = Workbook(write_only=True)
# ws = wb.create_sheet()
# cell = WriteOnlyCell(ws)
# cell.style = 'Pandas'

# def format_first_row(row, cell):
#     for c in row:
#         cell.value = c
#         yield cell

# rows = dataframe_to_rows(ratiosDF)
# print(ratiosDF.row)

# first_row = format_first_row(next(rows), cell)
# ws.append(first_row)

# for row in rows:
#     row = list(row)
#     cell.value = row[0]
#     row[0] = cell
#     ws.append(row)

# wb.save("openpyxl_stream.xlsx")

# print(ratiosDF)